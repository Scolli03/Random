import sys

print(sys.version)
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
import win32com.client
from configparser import ConfigParser

timestamp = datetime.datetime.today().strftime('%#m/%#d/%Y %I:%M')
cfg = ConfigParser()
cfg.read(r"\\dc01\atos_programs\MCXL\Templates\2018\gom_scripts\Config_Files\Production_Toolkit_Config.ini")
part = cfg.get('Part', 'name')
data = cfg.get(part, 'mold_master')
print(data)
template = cfg.get(part, 'template')
print(template)
sheet_name = cfg.get(part, 'sheet')
print(sheet_name)
DIR = cfg.get(part, 'mold_folder')
print(DIR)
wb = load_workbook(data, data_only=True, read_only=False, keep_vba=True)
wb2 = load_workbook(template, read_only=False, keep_vba=True)
sheet = wb.get_sheet_by_name(sheet_name)
sheet2 = wb2.get_sheet_by_name('Data')
if sheet_name == 'GE2679M19':
    start_row = 6
    start_column = 8
elif sheet_name == 'GE2725M07':
    start_row = 6
    start_column = 8
elif sheet_name == 'GE3121M70':
    start_row = 5
    start_column = 8
elif sheet_name == 'GE3121M73':
    start_row = 5
    start_column = 8
elif sheet_name == 'GE2553M90':
    start_row = 4
    start_column = 5
elif sheet_name == 'GE2550M30':
    start_row = 4
    start_column = 5
elif sheet_name == 'GE2706M84':
    start_row = 6
    start_column = 9

print(start_row)
print(start_column)
count = 1
check = True
while check != False:
    if sheet.cell(row=(start_row + 1) + count, column=start_column).value != None:
        count += 1
    else:
        check = False

count += start_row

sn = "start"
while sn != "null":
    sn = sheet.cell(row=start_row, column=start_column).value
    if sn != "null":
        letter = get_column_letter(start_column)
        data_range = (letter + str(start_row + 1) + ":" + letter + str((start_row + 1) + count))
        values = []

        for row in sheet.iter_rows(min_row=start_row + 1, max_row=count, min_col=start_column, max_col=start_column):
            for cell in row:
                if cell.value is not None and cell.value != 'null':
                    values.append(cell.value)
                else:
                    break
        r = 2
        for x in values:
            sheet2.cell(row=r, column=4).value = sn
            sheet2.cell(row=r, column=5).value = timestamp
            sheet2.cell(row=r, column=6).value = x
            r += 1
        path = (DIR + "\\" + str(sn) + ".xlsm")
        wb2.save(DIR + "\\" + str(sn) + ".xlsm")

        if os.path.exists(path):
            x1 = win32com.client.Dispatch("Excel.Application")
            x1.Workbooks.Open(Filename=path)
            x1.Application.Run("Button1_Click")
            x1.Application.Quit()
            del x1

        start_column += 1

    else:
        break
